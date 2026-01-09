import streamlit as st
import pdfplumber
import re
import requests
from openpyxl import load_workbook
from io import BytesIO

# ---------------------------
# BRØNNØYSUND LOOKUP FUNCTIONS
# ---------------------------
def lookup_by_org_number(org_number):
    url = f"https://data.brreg.no/enhetsregisteret/api/enheter/{org_number}"
    r = requests.get(url)
    if r.status_code == 200:
        return r.json()
    return None

def search_company_by_name(name):
    url = "https://data.brreg.no/enhetsregisteret/api/enheter"
    params = {"navn": name}
    r = requests.get(url, params=params)
    if r.status_code == 200:
        data = r.json()
        if "_embedded" in data and data["_embedded"].get("enheter"):
            return data["_embedded"]["enheter"][0]
    return None

# ---------------------------
# STREAMLIT SETUP
# ---------------------------
st.set_page_config(page_title="PDF → Excel Mapper", layout="centered")
st.title("📄➡️📊 PDF to Pre-Made Excel Template")

# ---------------------------
# PDF TEXT EXTRACTION
# ---------------------------
def extract_pdf_text(pdf_file):
    text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text

# ---------------------------
# FIND COMPANY IN PDF
# ---------------------------
def extract_company_from_pdf(text):
    fields = {}

    patterns = {
        "company_name": r"Company Name[:\s]+(.+)",
        "org_number": r"Org(?:anisation)? Number[:\s]+([\d\-]+)",
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        fields[key] = match.group(1).strip() if match else ""

    return fields

# ---------------------------
# EXCEL UPDATE
# ---------------------------
def update_excel(template_file, data, company_summary):
    wb = load_workbook(template_file)
    ws = wb.active

    cell_mapping = {
        "company_name": "B14",
        "org_number": "B15",
        "address": "B16",
        "post_nr": "B17",
        "nace_code": "B18",
        "turnover_2024": "B19",
        "homepage": "B21",
        "employees": "B22",
    }

    for field, cell in cell_mapping.items():
        if data.get(field):
            ws[cell] = data[field]

    if company_summary:
        ws["B10"] = f"Kort info om företaget:\n{company_summary}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ---------------------------
# UI
# ---------------------------
pdf_file = st.file_uploader("Upload PDF", type="pdf")
excel_file = st.file_uploader("Upload Excel Template", type=["xlsx"])

if pdf_file and excel_file:
    if st.button("Extract & Update Excel"):
        with st.spinner("Processing..."):

            # STEP 1 — PDF
            pdf_text = extract_pdf_text(pdf_file)
            extracted_fields = extract_company_from_pdf(pdf_text)

            company_name = extracted_fields.get("company_name", "")
            org_number = extracted_fields.get("org_number", "")

            # STEP 2 — BRØNNØYSUND
            company_data = None

            if org_number:
                company_data = lookup_by_org_number(org_number)

            if not company_data and company_name:
                company_data = search_company_by_name(company_name)

            company_summary = ""

            if company_data:
                extracted_fields["company_name"] = company_data.get("navn", "")
                extracted_fields["org_number"] = company_data.get("organisasjonsnummer", "")

                addr = company_data.get("forretningsadresse") or {}
                extracted_fields["address"] = " ".join(addr.get("adresse", []))
                extracted_fields["post_nr"] = addr.get("postnummer", "")

                nace = company_data.get("naeringskode1", {})
                extracted_fields["nace_code"] = nace.get("kode", "")

                extracted_fields["homepage"] = company_data.get("hjemmeside", "")
                extracted_fields["employees"] = company_data.get("antallAnsatte", "")

                # Simple official summary
                summary_parts = []
                if company_data.get("navn"):
                    summary_parts.append(company_data["navn"])
                if nace.get("beskrivelse"):
                    summary_parts.append(nace["beskrivelse"])

                company_summary = " – ".join(summary_parts)

            # STEP 3 — EXCEL
            updated_excel = update_excel(
                excel_file,
                extracted_fields,
                company_summary
            )

        st.success("Excel updated successfully!")
        st.subheader("Extracted data")
        st.json(extracted_fields)

        st.download_button(
            "Download updated Excel file",
            updated_excel,
            "updated_template.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
